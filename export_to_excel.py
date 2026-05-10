import openpyxl
import psycopg2
from copy import copy

TEMPLATE = "/Users/erikomaraliev/Documents/Проект Афаси/БД/Проект Айгуль Мырзабековна/Для Айг.FINISH RUS AFCD а веб сайт КР 25 03 2025(07.05.2026).xlsx"
OUTPUT = "/Users/erikomaraliev/IdeaProjects/afaci_back/export_afaci.xlsx"

DB_CONFIG = {
    "host": "localhost", "port": 5432,
    "dbname": "aafaci", "user": "erikomaraliev", "password": ""
}

conn = psycopg2.connect(**DB_CONFIG)
cur = conn.cursor()

def q(sql, params=()):
    cur.execute(sql, params)
    return cur.fetchall()

def get_val(pid, table, name_col, name_val, val_cols=("quantity", "error"), extra_where=""):
    cols = ", ".join(val_cols)
    sql = f"SELECT {cols} FROM {table} WHERE product_id = %s AND {name_col} ILIKE %s {extra_where} LIMIT 1"
    cur.execute(sql, (pid, name_val))
    r = cur.fetchone()
    if r and any(v is not None for v in r):
        parts = []
        for v in r:
            if v is not None:
                parts.append(str(v).replace(".", ","))
            else:
                parts.append("")
        if parts[1]:
            return f"{parts[0]}±{parts[1]}"
        return parts[0]
    return ""

def fmt_val(q, e):
    if q is None and e is None:
        return ""
    qs = str(q).replace(".", ",") if q is not None else ""
    es = str(e).replace(".", ",") if e is not None else ""
    if es and float(e) > 0:
        return f"{qs}±{es}"
    return qs

AMINO_ACIDS = [
    "Аспарагиновая кислота", "Глутаминовая кислота", "Серин", "Гистидин",
    "Глицин", "Треонин", "Аргинин", "Аланин", "Тирозин", "Цистин",
    "Валин", "Метионин", "Триптофан", "Фенилаланин", "Изолейцин",
    "Лейцин", "Лизин", "Пролин", "Аспарагин", "Глутамин"
]

# Build data cache: product_id -> { compound_name -> (q, e) }
print("Loading data from database...")

# Products by region
region_ids = {}
for r in q("SELECT id, name FROM regions"):
    region_ids[r[1].strip().lower()] = r[0]

# Products
products_by_region = {}
for db_region_name in ["Чуйская область", "Иссык-Кульская область", "Нарынская область", "Баткенская область", "Таласская область"]:
    rid = region_ids.get(db_region_name.strip().lower())
    if rid:
        rows = q("""
            SELECT p.id, p.name, c.name AS category
            FROM products p JOIN categories c ON c.id = p.categories_id
            WHERE p.regions_id = %s ORDER BY c.name, p.name
        """, (rid,))
        products_by_region[db_region_name] = rows

# Chemical composition cache
print("Caching chemical composition...")
chem_cache = {}
for row in q("SELECT product_id, compound_name, quantity, error, unit FROM chemical_composition"):
    pid = row[0]
    if pid not in chem_cache:
        chem_cache[pid] = {}
    chem_cache[pid][row[1].strip().lower()] = (row[2], row[3], row[4])

# Mineral composition cache
print("Caching minerals...")
min_cache = {}
for row in q("SELECT product_id, mineral_name, quantity, error FROM mineral_composition"):
    pid = row[0]
    if pid not in min_cache:
        min_cache[pid] = {}
    min_cache[pid][row[1].strip().lower()] = (row[2], row[3])

# Vitamin composition cache
print("Caching vitamins...")
vit_cache = {}
for row in q("SELECT product_id, vitamin_name, quantity, error FROM vitamin_composition"):
    pid = row[0]
    if pid not in vit_cache:
        vit_cache[pid] = {}
    vit_cache[pid][row[1].strip().lower()] = (row[2], row[3])

# Amino acid cache
print("Caching amino acids...")
aa_cache = {}
for row in q("SELECT product_id, amino_acid_name, quantity, error FROM amino_acid_composition"):
    pid = row[0]
    if pid not in aa_cache:
        aa_cache[pid] = {}
    aa_cache[pid][row[1].strip().lower()] = (row[2], row[3])

# Fatty acid cache
print("Caching fatty acids...")
fa_cache = {}
for row in q("SELECT product_id, fatty_acid_name, type_of_fatty_acid, quantity, error FROM fatty_acid_composition"):
    pid = row[0]
    if pid not in fa_cache:
        fa_cache[pid] = {}
    fa_type = row[2].strip().lower() if row[2] else ""
    fa_cache[pid][(row[1].strip().lower(), fa_type)] = (row[3], row[4])

# Kyrgyz names
kg_cache = {}
for row in q("SELECT product_id, product_name FROM products_translate WHERE language = 'KG'"):
    kg_cache[row[0]] = row[1]

def get_chem(pid, name):
    d = chem_cache.get(pid, {})
    r = d.get(name.strip().lower())
    if r:
        return fmt_val(r[0], r[1])
    return ""

def get_min(pid, name):
    d = min_cache.get(pid, {})
    r = d.get(name.strip().lower())
    if r:
        return fmt_val(r[0], r[1])
    return ""

def get_vit(pid, name):
    d = vit_cache.get(pid, {})
    r = d.get(name.strip().lower())
    if r:
        return fmt_val(r[0], r[1])
    return ""

def get_aa(pid, name):
    d = aa_cache.get(pid, {})
    r = d.get(name.strip().lower())
    if r:
        return fmt_val(r[0], r[1])
    return ""

def get_fa(pid, name, fa_type):
    d = fa_cache.get(pid, {})
    r = d.get((name.strip().lower(), fa_type.strip().lower()))
    if r:
        return fmt_val(r[0], r[1])
    return ""

def get_energy(pid):
    """Get energy as kcal/kJ combined string"""
    kcal = chem_cache.get(pid, {}).get("энерг. ценность")
    if kcal and kcal[2] == 'kcal':
        # Find kJ version
        for k, v in chem_cache.get(pid, {}).items():
            if "ценность" in k and v[2] == 'kJ':
                return f"{str(kcal[0]).replace('.', ',')}/{str(v[0]).replace('.', ',')} кДж"
        return str(kcal[0]).replace(".", ",")
    return ""

def get_chem_like(pid, pattern):
    for k, v in chem_cache.get(pid, {}).items():
        if k.startswith(pattern.lower()) or pattern.lower() in k:
            return fmt_val(v[0], v[1])
    return ""

# Region name mapping (sheet name -> DB region name)
SHEET_REGION = {
    "Чуйская область ": "Чуйская область",
    "Иссык-Кульская область": "Иссык-Кульская область",
    "Нарынская область": "Нарынская область",
    "Баткенская область": "Баткенская область",
    "Таласская область": "Таласская область",
}

print("Loading template...")
wb = openpyxl.load_workbook(TEMPLATE)

for sheet_name, db_region in SHEET_REGION.items():
    ws = wb[sheet_name]
    products = products_by_region.get(db_region, [])
    print(f"\n{db_region}: {len(products)} products")

    # Read headers (row 3) to map columns
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(3, c).value
        if v:
            headers[c] = str(v).strip().replace('\n', ' ')

    # Find data start row
    data_start = None
    for r in range(1, 10):
        v = ws.cell(r, 1).value
        if v and str(v).strip().isdigit():
            data_start = r
            break
    if data_start is None:
        data_start = 7

    # Map columns by header name
    col_idx = 1
    col_category = None
    col_product = 3  # col 3 is always Продукт и описание
    col_product_kg = 4  # col 4 is always Кыргызское название
    col_energy = 5
    col_moisture = 6
    col_protein = 7
    col_fat = 8
    col_ash = 9
    col_carbs = 10
    col_ca = 11 if ws.cell(3, 11).value and "Кальций" in str(ws.cell(3, 11).value) else None
    col_fe = None
    col_p = None
    col_k = None
    col_na = None
    col_vit_a = None
    col_vit_b1 = None
    col_vit_b2 = None
    col_vit_c = None

    # Scan headers for column positions
    aa_columns = {}
    fa_saturated_columns = {}
    fa_mono_columns = {}
    fa_poly_columns = {}
    ratio_columns = {}

    # Build section headers from row 2 to know column groups
    section_map = {}  # column -> section name
    current_section = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v:
            sv = str(v).strip().rstrip(':')
            if sv in ("Насыщенные жирные кислоты", "Мононенасыщенные жирные кислоты",
                      "Полиненасыщенные жирные кислоты"):
                current_section = sv
            elif "Соотношения" in sv:
                current_section = "Соотношения"
            elif sv in ("средние значения", "Минералы", "Витамины", "Аминокислотный состав", ""):
                current_section = None
            else:
                current_section = None
        section_map[c] = current_section

    for c, h in headers.items():
        h_clean = h.strip()
        if h_clean == "Группа продукта":
            col_category = c
        elif h_clean == "Продукт и описание":
            col_product = c
        elif h_clean.startswith("Нимаенование"):
            col_product_kg = c
        elif h_clean == "Энергетическая ценность":
            col_energy = c
        elif h_clean == "Влажность":
            col_moisture = c
        elif h_clean == "Белок":
            col_protein = c
        elif h_clean == "Жир":
            col_fat = c
        elif h_clean == "Зольность":
            col_ash = c
        elif h_clean == "Углеводы":
            col_carbs = c
        elif h_clean == "Кальций":
            col_ca = c
        elif h_clean == "Железо":
            col_fe = c
        elif h_clean == "Фосфор":
            col_p = c
        elif h_clean in ("Калий", " Калий"):
            col_k = c
        elif h_clean == "Натрий":
            col_na = c
        elif h_clean == "Витамин А":
            col_vit_a = c
        elif h_clean == "Витамин В1":
            col_vit_b1 = c
        elif h_clean == "Витамин В2":
            col_vit_b2 = c
        elif h_clean == "Витамин С":
            col_vit_c = c
        elif h_clean in AMINO_ACIDS:
            aa_columns[h_clean] = c
        elif section_map.get(c) == "Насыщенные жирные кислоты":
            fa_saturated_columns[h_clean] = c
        elif section_map.get(c) == "Мононенасыщенные жирные кислоты":
            fa_mono_columns[h_clean] = c
        elif section_map.get(c) == "Полиненасыщенные жирные кислоты":
            fa_poly_columns[h_clean] = c
        elif section_map.get(c) == "Соотношения":
            ratio_columns[h_clean] = c

    # Clear data rows (skip merged cells)
    merged = set()
    for mr in ws.merged_cells.ranges:
        for mrow in range(mr.min_row, mr.max_row + 1):
            for mcol in range(mr.min_col, mr.max_col + 1):
                merged.add((mrow, mcol))
    for r in range(data_start, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if (r, c) not in merged:
                ws.cell(r, c).value = None

    # Fill data
    row_idx = data_start
    for idx, (pid, pname, pcat) in enumerate(products, 1):
        ws.cell(row_idx, 1, idx)
        if col_category:
            ws.cell(row_idx, col_category, pcat)
        ws.cell(row_idx, col_product, pname)
        ws.cell(row_idx, col_product_kg, kg_cache.get(pid, ""))

        # Energy
        ws.cell(row_idx, col_energy, get_energy(pid))

        # Chemical
        if col_moisture:
            ws.cell(row_idx, col_moisture, get_chem_like(pid, "влажн"))
        if col_protein:
            ws.cell(row_idx, col_protein, get_chem_like(pid, "белок"))
        if col_fat:
            ws.cell(row_idx, col_fat, get_chem_like(pid, "жир"))
        if col_ash:
            ws.cell(row_idx, col_ash, get_chem_like(pid, "зольн"))
        if col_carbs:
            ws.cell(row_idx, col_carbs, get_chem_like(pid, "углев"))

        # Minerals
        if col_ca: ws.cell(row_idx, col_ca, get_min(pid, "Ca"))
        if col_fe: ws.cell(row_idx, col_fe, get_min(pid, "Fe"))
        if col_p: ws.cell(row_idx, col_p, get_min(pid, "P"))
        if col_k: ws.cell(row_idx, col_k, get_min(pid, "K"))
        if col_na: ws.cell(row_idx, col_na, get_min(pid, "Na"))

        # Vitamins
        if col_vit_a: ws.cell(row_idx, col_vit_a, get_vit(pid, "Витамин А"))
        if col_vit_b1: ws.cell(row_idx, col_vit_b1, get_vit(pid, "Витамин В1"))
        if col_vit_b2: ws.cell(row_idx, col_vit_b2, get_vit(pid, "Витамин В2"))
        if col_vit_c: ws.cell(row_idx, col_vit_c, get_vit(pid, "Витамин С"))

        # Amino acids
        for aa_name, col in aa_columns.items():
            val = get_aa(pid, aa_name)
            if val:
                ws.cell(row_idx, col, val)

        # Saturated fatty acids
        for fa_name, col in fa_saturated_columns.items():
            val = get_fa(pid, fa_name, "Насыщенные ЖК")
            if val:
                ws.cell(row_idx, col, val)

        # Monounsaturated fatty acids
        for fa_name, col in fa_mono_columns.items():
            val = get_fa(pid, fa_name, "Мононенасыщенные ЖК")
            if val:
                ws.cell(row_idx, col, val)

        # Polyunsaturated fatty acids
        for fa_name, col in fa_poly_columns.items():
            val = get_fa(pid, fa_name, "Полиненасыщенные ЖК")
            if val:
                ws.cell(row_idx, col, val)

        # Ratios
        for ratio_name, col in ratio_columns.items():
            val = get_fa(pid, ratio_name, "Соотношения метиловых эфиров жирных кислот молочного жира")
            if val:
                ws.cell(row_idx, col, val)

        row_idx += 1

    print(f"  Written {idx} rows")

print(f"\nSaving to {OUTPUT}...")
wb.save(OUTPUT)
print("Done!")
cur.close()
conn.close()
